import streamlit as st
import sqlite3
import json
import openpyxl
from io import BytesIO

#  CONFIG & PATHS 
DB_PATH = 'edtech_eval.db'
REQ_FILE = "Requirements.xlsx"

st.set_page_config(page_title="Scoring Matrix", layout="wide")
st.title("Scoring Matrix")

def load_excel_requirements(file_path):
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        rows = list(sheet.iter_rows(values_only=True))
        header = rows[0]
        return [dict(zip(header, row)) for row in rows[1:] if row[0] or row[1]]
    except Exception as e:
        st.error(f"Error loading Excel: {e}")
        return []

def convert_to_excel(header, data):
    """Creates an Excel file in memory using openpyxl."""
    output = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Master Matrix"
    
    # Write Header
    ws.append(header)
    
    # Write Data Rows
    for row in data:
        ws.append(row)
        
    # Formatting
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        
    wb.save(output)
    return output.getvalue()

#  DATA LOADING 
conn = sqlite3.connect(DB_PATH)
try:
    cursor = conn.execute("SELECT app_name, scores FROM evaluations")
    db_rows = cursor.fetchall()
except sqlite3.OperationalError:
    st.warning("Database not found. Please evaluate platforms on the main page first.")
    st.stop()
finally:
    conn.close()

req_list = load_excel_requirements(REQ_FILE)

if not db_rows or not req_list:
    st.warning("No data available to display.")
    st.stop()

#  SIDEBAR FILTERS 
st.sidebar.header("Matrix Controls")

all_available_apps = [row[0] for row in db_rows]
selected_apps = st.sidebar.multiselect(
    "Select Apps to Compare", 
    options=all_available_apps, 
    default=all_available_apps
)

all_cats = sorted(list(set(r['Category'] for r in req_list if r.get('Category'))))
selected_cat = st.sidebar.multiselect("Filter by Category", all_cats, default=all_cats)

#  BUILDING THE MATRIX 
header_base = ["ID", "Description", "Priority"]
full_header = header_base + selected_apps
matrix_data = []

filtered_db_rows = [row for row in db_rows if row[0] in selected_apps]

for req in req_list:
    if req.get('Category') not in selected_cat:
        continue
        
    rid = req.get('Req ID')
    desc = req.get('Requirement Description')
    priority = req.get('Priority', 'N/A')
    
    row_data = [rid, desc, priority]
    
    for app_name, score_json in filtered_db_rows:
        app_scores = json.loads(score_json)
        score = app_scores.get(rid, "-")
        row_data.append(score)
        
    matrix_data.append(row_data)

#  UI DISPLAY 
st.subheader("Score Comparison")

if selected_apps:
    excel_data = convert_to_excel(full_header, matrix_data)
    
    st.download_button(
        label="Download Matrix",
        data=excel_data,
        file_name="EdTech_Evaluation_Matrix.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary"
    )

    # Display Table
    table_ready_data = []
    for r in matrix_data:
        table_ready_data.append(dict(zip(full_header, r)))
    
    st.table(table_ready_data)
else:
    st.info("Please select at least one app in the sidebar to view the matrix.")